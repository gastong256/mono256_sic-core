import pytest

try:
    import openpyxl

    OPENPYXL_AVAILABLE = True
except ModuleNotFoundError:
    OPENPYXL_AVAILABLE = False

from apps.reports.exports import (
    build_journal_book_workbook,
    build_ledger_workbook,
    build_trial_balance_workbook,
)


def _active_sheet(artifact):
    artifact.workbook.close()
    artifact.stream.seek(0)
    loaded = openpyxl.load_workbook(artifact.stream)
    return loaded.active


@pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl is not installed")
class TestReportExcelExports:
    def test_journal_book_workbook_has_expected_headers(self):
        report = {
            "company": "ACME",
            "date_from": "2026-01-01",
            "date_to": "2026-01-31",
            "entries": [
                {
                    "entry_number": 1,
                    "date": "2026-01-10",
                    "description": "Venta",
                    "source_ref": "FAC-1",
                    "lines": [
                        {
                            "account_code": "1.01.01",
                            "account_name": "Caja",
                            "debit": "100.00",
                            "credit": None,
                        },
                        {
                            "account_code": "5.01.01",
                            "account_name": "Ventas",
                            "debit": None,
                            "credit": "100.00",
                        },
                    ],
                    "total_debit": "100.00",
                    "total_credit": "100.00",
                }
            ],
            "totals": {"total_debit": "100.00", "total_credit": "100.00"},
        }

        artifact = build_journal_book_workbook(report=report)
        ws = _active_sheet(artifact)

        assert ws.title == "Libro Diario"
        assert ws["A1"].value == "Libro Diario"
        assert ws["A4"].value == "Asiento"

    def test_ledger_workbook_has_account_sections(self):
        report = {
            "company": "ACME",
            "date_from": "2026-01-01",
            "date_to": "2026-01-31",
            "accounts": [
                {
                    "account_code": "1.01.01",
                    "account_name": "Caja",
                    "account_type": "AS",
                    "normal_balance": "DEBIT",
                    "opening_balance": "0.00",
                    "movements": [],
                    "period_totals": {"total_debit": "100.00", "total_credit": "0.00"},
                    "closing_balance": "100.00",
                }
            ],
        }

        artifact = build_ledger_workbook(report=report)
        ws = _active_sheet(artifact)

        assert ws.title == "Libro Mayor"
        assert ws["A1"].value == "Libro Mayor"
        assert "1.01.01 - Caja" in str(ws["A5"].value)

    def test_trial_balance_workbook_has_group_and_totals(self):
        report = {
            "company": "ACME",
            "date_from": "2026-01-01",
            "date_to": "2026-01-31",
            "groups": [
                {
                    "account_code": "1.01",
                    "account_name": "Caja",
                    "account_type": "AS",
                    "subtotal_debit": "100.00",
                    "subtotal_credit": "0.00",
                    "subtotal_debit_balance": "100.00",
                    "subtotal_credit_balance": None,
                    "accounts": [
                        {
                            "account_code": "1.01.01",
                            "account_name": "Caja chica",
                            "account_type": "AS",
                            "total_debit": "100.00",
                            "total_credit": "0.00",
                            "debit_balance": "100.00",
                            "credit_balance": None,
                        }
                    ],
                }
            ],
            "totals": {
                "total_debit": "100.00",
                "total_credit": "0.00",
                "total_debit_balance": "100.00",
                "total_credit_balance": "0.00",
            },
        }

        artifact = build_trial_balance_workbook(report=report)
        ws = _active_sheet(artifact)

        assert ws.title == "Balance"
        assert ws["A1"].value == "Balance de Comprobacion"
        assert ws["B4"].value == "Cuenta"
