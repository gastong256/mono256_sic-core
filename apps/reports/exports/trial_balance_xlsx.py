from __future__ import annotations

def build_trial_balance_workbook(*, report: dict) -> object:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Balance"

    bold = Font(bold=True)

    ws["A1"] = "Balance de Comprobacion"
    ws["A1"].font = bold
    ws["A2"] = f"Empresa: {report['company']}"
    ws["A3"] = f"Desde: {report['date_from']}  Hasta: {report['date_to']}"

    headers = [
        "Codigo",
        "Cuenta",
        "Tipo",
        "Debe",
        "Haber",
        "Saldo Deudor",
        "Saldo Acreedor",
    ]
    ws.append(headers)
    for cell in ws[4]:
        cell.font = bold

    for group in report.get("groups", []):
        ws.append(
            [
                group.get("account_code"),
                group.get("account_name"),
                group.get("account_type"),
                group.get("subtotal_debit"),
                group.get("subtotal_credit"),
                group.get("subtotal_debit_balance"),
                group.get("subtotal_credit_balance"),
            ]
        )
        for cell in ws[ws.max_row]:
            cell.font = bold

        for account in group.get("accounts", []):
            ws.append(
                [
                    account.get("account_code"),
                    account.get("account_name"),
                    account.get("account_type"),
                    account.get("total_debit"),
                    account.get("total_credit"),
                    account.get("debit_balance"),
                    account.get("credit_balance"),
                ]
            )

    ws.append([])
    totals = report.get("totals", {})
    ws.append(
        [
            "",
            "Totales",
            "",
            totals.get("total_debit", "0.00"),
            totals.get("total_credit", "0.00"),
            totals.get("total_debit_balance", "0.00"),
            totals.get("total_credit_balance", "0.00"),
        ]
    )
    for cell in ws[ws.max_row]:
        cell.font = bold

    return wb
