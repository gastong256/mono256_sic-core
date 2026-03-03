from __future__ import annotations

def build_ledger_workbook(*, report: dict) -> object:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Libro Mayor"

    bold = Font(bold=True)

    ws["A1"] = "Libro Mayor"
    ws["A1"].font = bold
    ws["A2"] = f"Empresa: {report['company']}"
    ws["A3"] = f"Desde: {report['date_from']}  Hasta: {report['date_to']}"
    ws.append([])

    for account in report.get("accounts", []):
        ws.append(
            [
                f"{account.get('account_code')} - {account.get('account_name')}",
                f"Tipo: {account.get('account_type')}",
                f"Saldo normal: {account.get('normal_balance')}",
            ]
        )
        for cell in ws[ws.max_row]:
            cell.font = bold

        ws.append(["Saldo inicial", "", "", "", "", account.get("opening_balance")])
        headers = ["Fecha", "Asiento", "Descripcion", "Referencia", "Debe", "Haber", "Saldo"]
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.font = bold

        for move in account.get("movements", []):
            ws.append(
                [
                    move.get("date"),
                    move.get("entry_number"),
                    move.get("description"),
                    move.get("source_ref"),
                    move.get("debit"),
                    move.get("credit"),
                    move.get("balance"),
                ]
            )

        ws.append(
            [
                "",
                "",
                "Totales periodo",
                "",
                account.get("period_totals", {}).get("total_debit", "0.00"),
                account.get("period_totals", {}).get("total_credit", "0.00"),
                account.get("closing_balance"),
            ]
        )
        for cell in ws[ws.max_row]:
            cell.font = bold
        ws.append([])

    return wb
