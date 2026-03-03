from __future__ import annotations

def build_journal_book_workbook(*, report: dict) -> object:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Libro Diario"

    bold = Font(bold=True)

    ws["A1"] = "Libro Diario"
    ws["A1"].font = bold
    ws["A2"] = f"Empresa: {report['company']}"
    ws["A3"] = f"Desde: {report['date_from']}  Hasta: {report['date_to']}"

    headers = ["Asiento", "Fecha", "Descripcion", "Referencia", "Cuenta", "Debe", "Haber"]
    ws.append(headers)
    for cell in ws[4]:
        cell.font = bold

    for entry in report.get("entries", []):
        for line in entry.get("lines", []):
            ws.append(
                [
                    entry.get("entry_number"),
                    entry.get("date"),
                    entry.get("description"),
                    entry.get("source_ref"),
                    f"{line.get('account_code')} - {line.get('account_name')}",
                    line.get("debit"),
                    line.get("credit"),
                ]
            )
        ws.append(
            [
                "",
                "",
                "Subtotal asiento",
                "",
                "",
                entry.get("total_debit"),
                entry.get("total_credit"),
            ]
        )
        for cell in ws[ws.max_row]:
            cell.font = bold
        ws.append([])

    ws.append(
        [
            "",
            "",
            "Totales",
            "",
            "",
            report.get("totals", {}).get("total_debit", "0.00"),
            report.get("totals", {}).get("total_credit", "0.00"),
        ]
    )
    for cell in ws[ws.max_row]:
        cell.font = bold

    return wb
